"""基于 roster 生成参赛者名单工作簿。"""

from __future__ import annotations

import re
from collections.abc import Callable
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from math import ceil
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from pubg_match_analyzer.core.constants import (
    display_game_mode,
    display_game_mode_category,
    to_display_team_no,
)
from pubg_match_analyzer.core.models import MatchOverview, TeamSummary
from pubg_match_analyzer.services.match_details import build_match_overview, extract_team_summaries
from pubg_match_analyzer.services.pubg_api import PubgAPIClient
from pubg_match_analyzer.services.signup_mapping import (
    ContactResolution,
    SignupContactLookup,
    SignupSheetSchema,
)

THIN_SIDE = Side(style="thin", color="000000")
ALL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
TITLE_FILL = PatternFill(fill_type="solid", fgColor="0D0D0D")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="0D0D0D")
INDEX_FILL = PatternFill(fill_type="solid", fgColor="262626")
MISSING_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
CONFLICT_FILL = PatternFill(fill_type="solid", fgColor="F4B183")
WHITE_FONT = Font(name="Microsoft YaHei", color="FFFFFF", bold=True)
TITLE_FONT = Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=16)
DATA_FONT = Font(name="Microsoft YaHei", color="000000")
CENTER = Alignment(horizontal="center", vertical="center")


@dataclass
class TeamParticipantGroup:
    """用于渲染名单模板的队伍数据。"""

    display_team_no: int
    players: list[ContactResolution]


@dataclass
class ParticipantListResult:
    """参赛者名单生成结果。"""

    workbook_bytes: bytes
    template_type: str
    signup_mode_name: str
    total_players: int
    filled_qq_count: int
    missing_contact_count: int
    conflict_count: int
    used_signup_sheet: bool


@dataclass
class BatchParticipantListResult:
    """批量参赛名单 ZIP 生成结果。"""

    zip_bytes: bytes
    zip_filename: str
    requested_match_count: int
    generated_match_count: int
    total_players: int
    total_conflicts: int
    total_missing_contacts: int
    item_filenames: list[str]
    failed_matches: list[dict[str, str]]


def _int_or_none(value: object) -> int | None:
    if value is None or value == "" or isinstance(value, bool):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _participant_team_no(team: TeamSummary) -> int:
    return int(to_display_team_no(_int_or_none(team.source_team_id), team.team_index) or team.team_index)


def _participant_team_sort_key(team: TeamSummary) -> tuple[int, int, int]:
    source_team_id = _int_or_none(team.source_team_id)
    if source_team_id is None:
        return (1, team.team_index, team.team_index)
    return (0, source_team_id, team.team_index)


def infer_participant_template(teams: list[TeamSummary]) -> tuple[str, str]:
    """根据队伍人数结构推导名单模板类型。"""
    if not teams:
        return "四排模板", "squad"

    counts = [team.player_count for team in teams]
    if all(count == 1 for count in counts):
        return "单排模板", "solo"
    if max(counts) > 4:
        return "1队N人模板", "multi"
    return "四排模板", "squad"


def build_participant_list_workbook(
    *,
    overview: MatchOverview,
    teams: list[TeamSummary],
    signup_excel_bytes: bytes | None = None,
    signup_sheet_schema: SignupSheetSchema | None = None,
    signup_mode_name: str | None = None,
    event_name: str | None = None,
    round_name: str | None = None,
) -> ParticipantListResult:
    """生成参赛者名单工作簿。"""
    template_type_label, template_type_code = infer_participant_template(teams)
    lookup = None
    if signup_excel_bytes and signup_sheet_schema is not None:
        lookup = SignupContactLookup.from_excel_bytes(signup_excel_bytes, signup_sheet_schema)

    groups: list[TeamParticipantGroup] = []
    total_players = 0
    filled_qq_count = 0
    missing_contact_count = 0
    conflict_rows: list[dict[str, str]] = []

    for team in sorted(teams, key=_participant_team_sort_key):
        players: list[ContactResolution] = []
        for player_name in team.player_names:
            total_players += 1
            if lookup is None:
                result = ContactResolution(player_name=player_name, status="no_signup")
            else:
                result = lookup.resolve(player_name, signup_mode_name)

            if result.qq:
                filled_qq_count += 1
            if result.status == "missing":
                missing_contact_count += 1
            if result.status == "conflict":
                conflict_rows.append(
                    {
                        "对局ID": overview.match_id,
                        "玩家昵称": player_name,
                        "候选QQ": " / ".join(result.candidate_qqs),
                        "采用QQ": result.qq,
                        "冲突来源条数": str(len(result.candidate_qqs)),
                        "最新提交时间": result.latest_submitted_at,
                    }
                )
            players.append(result)
        groups.append(
            TeamParticipantGroup(
                display_team_no=_participant_team_no(team),
                players=players,
            )
        )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "参赛者名单"

    title = _build_sheet_title(overview, event_name=event_name, round_name=round_name)
    if template_type_code == "solo":
        _render_solo_sheet(worksheet, title, groups)
    elif template_type_code == "multi":
        _render_multi_team_sheet(worksheet, title, groups)
    else:
        _render_squad_sheet(worksheet, title, groups)

    if conflict_rows:
        _render_conflict_sheet(workbook.create_sheet("QQ冲突"), conflict_rows)

    buffer = BytesIO()
    workbook.save(buffer)
    return ParticipantListResult(
        workbook_bytes=buffer.getvalue(),
        template_type=template_type_label,
        signup_mode_name=signup_mode_name or "",
        total_players=total_players,
        filled_qq_count=filled_qq_count,
        missing_contact_count=missing_contact_count,
        conflict_count=len(conflict_rows),
        used_signup_sheet=signup_excel_bytes is not None and signup_sheet_schema is not None,
    )


def _build_sheet_title(
    overview: MatchOverview,
    event_name: str | None = None,
    round_name: str | None = None,
) -> str:
    """构造主工作表标题。"""
    category = display_game_mode_category(overview.custom_match_category)
    mode = display_game_mode(overview.game_mode)
    mode_text = " ".join(part for part in (category, mode if mode and mode != "-" else "") if part)
    if not mode_text:
        mode_text = overview.game_mode or "对局"
    event_text = sanitize_filename_part(event_name)
    round_text = sanitize_filename_part(round_name)
    if event_text and round_text:
        return f"{event_text} {round_text}"
    if event_text:
        return event_text
    if round_text:
        return f"{mode_text} {round_text}"
    return mode_text


def _style_cell(
    cell,
    *,
    value: Any = None,
    font: Font | None = None,
    fill: PatternFill | None = None,
    border: Border | None = None,
    alignment: Alignment | None = None,
) -> None:
    """统一设置单元格样式。"""
    cell.value = value
    if font is not None:
        cell.font = copy(font)
    if fill is not None:
        cell.fill = copy(fill)
    if border is not None:
        cell.border = copy(border)
    if alignment is not None:
        cell.alignment = copy(alignment)


def _apply_contact_fill(cell, status: str) -> None:
    """按匹配状态给 QQ 单元格着色。"""
    if status == "missing":
        cell.fill = copy(MISSING_FILL)
    elif status == "conflict":
        cell.fill = copy(CONFLICT_FILL)


def _write_qq_cell(cell, qq: str, status: str) -> None:
    """按 Excel 友好的方式写入 QQ，避免“数字储存成文字”提示。"""
    normalized = str(qq or "").strip()
    if normalized.isdigit() and not (len(normalized) > 1 and normalized.startswith("0")):
        cell.value = int(normalized)
        cell.number_format = "0"
    else:
        cell.value = normalized
        cell.number_format = "@"
    _apply_contact_fill(cell, status)


def _set_column_widths(ws: Worksheet, widths: dict[int, float]) -> None:
    """批量设置列宽。"""
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _prepare_title(ws: Worksheet, title: str, start_col: int, end_col: int) -> None:
    """写入顶部标题区。"""
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=start_col, end_row=5, end_column=end_col)
    cell = ws.cell(1, start_col)
    _style_cell(cell, value=title, font=TITLE_FONT, fill=TITLE_FILL, border=ALL_BORDER, alignment=CENTER)
    for row in range(1, 6):
        ws.row_dimensions[row].height = 24
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).border = copy(ALL_BORDER)
            if row > 1:
                ws.cell(row, col).fill = copy(TITLE_FILL)


def _render_squad_sheet(ws: Worksheet, title: str, groups: list[TeamParticipantGroup]) -> None:
    """按四排模板渲染名单。"""
    start_col = 3
    teams_per_band = 4
    end_col = start_col + teams_per_band * 2 - 1
    _prepare_title(ws, title, start_col, end_col)
    _set_column_widths(ws, {3: 13.0, 4: 10.5, 5: 13.0, 6: 10.5, 7: 13.0, 8: 10.5, 9: 13.0, 10: 10.5})

    row_base = 6
    band_height = 6
    for band_index in range(ceil(len(groups) / teams_per_band)):
        band = groups[band_index * teams_per_band : (band_index + 1) * teams_per_band]
        for offset in range(teams_per_band):
            col = start_col + offset * 2
            team = band[offset] if offset < len(band) else None
            ws.merge_cells(start_row=row_base, start_column=col, end_row=row_base, end_column=col + 1)
            _style_cell(
                ws.cell(row_base, col),
                value=f"TEAM# {team.display_team_no}" if team else "",
                font=WHITE_FONT,
                fill=HEADER_FILL,
                border=ALL_BORDER,
                alignment=CENTER,
            )
            for title_offset, header in enumerate(("游戏ID", "QQ")):
                _style_cell(
                    ws.cell(row_base + 1, col + title_offset),
                    value=header,
                    font=WHITE_FONT,
                    fill=HEADER_FILL,
                    border=ALL_BORDER,
                    alignment=CENTER,
                )
            for player_offset in range(4):
                row = row_base + 2 + player_offset
                player = team.players[player_offset] if team and player_offset < len(team.players) else None
                _style_cell(
                    ws.cell(row, col),
                    value=player.player_name if player else "",
                    font=DATA_FONT,
                    border=ALL_BORDER,
                    alignment=CENTER,
                )
                qq_cell = ws.cell(row, col + 1)
                _style_cell(
                    qq_cell,
                    value="",
                    font=DATA_FONT,
                    border=ALL_BORDER,
                    alignment=CENTER,
                )
                if player:
                    _write_qq_cell(qq_cell, player.qq, player.status)
        row_base += band_height


def _render_solo_sheet(ws: Worksheet, title: str, groups: list[TeamParticipantGroup]) -> None:
    """按单排模板渲染名单。"""
    start_col = 3
    blocks_per_row = 4
    end_col = start_col + blocks_per_row * 3 - 1
    _prepare_title(ws, title, start_col, end_col)
    _set_column_widths(
        ws,
        {
            3: 5.0,
            4: 14.0,
            5: 10.5,
            6: 5.0,
            7: 14.0,
            8: 10.5,
            9: 5.0,
            10: 14.0,
            11: 10.5,
            12: 5.0,
            13: 14.0,
            14: 10.5,
        },
    )

    row_base = 6
    for offset in range(blocks_per_row):
        col = start_col + offset * 3
        for title_offset, header in enumerate(("队伍", "游戏ID", "QQ")):
            _style_cell(
                ws.cell(row_base, col + title_offset),
                value=header,
                font=WHITE_FONT,
                fill=HEADER_FILL,
                border=ALL_BORDER,
                alignment=CENTER,
            )

    data_row = row_base + 1
    for group_index in range(ceil(len(groups) / blocks_per_row)):
        row = data_row + group_index
        band = groups[group_index * blocks_per_row : (group_index + 1) * blocks_per_row]
        for offset in range(blocks_per_row):
            col = start_col + offset * 3
            team = band[offset] if offset < len(band) else None
            player = team.players[0] if team and team.players else None
            _style_cell(
                ws.cell(row, col),
                value=team.display_team_no if team else "",
                font=WHITE_FONT,
                fill=INDEX_FILL,
                border=ALL_BORDER,
                alignment=CENTER,
            )
            _style_cell(
                ws.cell(row, col + 1),
                value=player.player_name if player else "",
                font=DATA_FONT,
                border=ALL_BORDER,
                alignment=CENTER,
            )
            qq_cell = ws.cell(row, col + 2)
            _style_cell(
                qq_cell,
                value="",
                font=DATA_FONT,
                border=ALL_BORDER,
                alignment=CENTER,
            )
            if player:
                _write_qq_cell(qq_cell, player.qq, player.status)


def _render_multi_team_sheet(ws: Worksheet, title: str, groups: list[TeamParticipantGroup]) -> None:
    """按 1 队 N 人模板渲染名单。"""
    start_col = 3
    teams_per_band = 3
    end_col = start_col + teams_per_band * 3 - 1
    _prepare_title(ws, title, start_col, end_col)
    _set_column_widths(ws, {3: 5.0, 4: 14.0, 5: 10.5, 6: 5.0, 7: 14.0, 8: 10.5, 9: 5.0, 10: 14.0, 11: 10.5})

    max_members = max((len(group.players) for group in groups), default=1)
    row_base = 6
    band_height = max_members + 3
    for band_index in range(ceil(len(groups) / teams_per_band)):
        band = groups[band_index * teams_per_band : (band_index + 1) * teams_per_band]
        for offset in range(teams_per_band):
            col = start_col + offset * 3
            team = band[offset] if offset < len(band) else None
            ws.merge_cells(start_row=row_base, start_column=col, end_row=row_base + 1, end_column=col + 2)
            _style_cell(
                ws.cell(row_base, col),
                value=f"TEAM# {team.display_team_no}" if team else "",
                font=WHITE_FONT,
                fill=HEADER_FILL,
                border=ALL_BORDER,
                alignment=CENTER,
            )
            for row in (row_base, row_base + 1):
                for current_col in range(col, col + 3):
                    ws.cell(row, current_col).border = copy(ALL_BORDER)
                    if row == row_base + 1:
                        ws.cell(row, current_col).fill = copy(HEADER_FILL)
            for title_offset, header in enumerate(("队内序号", "游戏ID", "QQ")):
                _style_cell(
                    ws.cell(row_base + 2, col + title_offset),
                    value=header,
                    font=WHITE_FONT,
                    fill=HEADER_FILL,
                    border=ALL_BORDER,
                    alignment=CENTER,
                )
            for player_offset in range(max_members):
                row = row_base + 3 + player_offset
                player = team.players[player_offset] if team and player_offset < len(team.players) else None
                index_cell = ws.cell(row, col)
                _style_cell(
                    index_cell,
                    value=player_offset + 1 if player else "",
                    font=WHITE_FONT if player else DATA_FONT,
                    fill=INDEX_FILL if player else None,
                    border=ALL_BORDER,
                    alignment=CENTER,
                )
                _style_cell(
                    ws.cell(row, col + 1),
                    value=player.player_name if player else "",
                    font=DATA_FONT,
                    border=ALL_BORDER,
                    alignment=CENTER,
                )
                qq_cell = ws.cell(row, col + 2)
                _style_cell(
                    qq_cell,
                    value="",
                    font=DATA_FONT,
                    border=ALL_BORDER,
                    alignment=CENTER,
                )
                if player:
                    _write_qq_cell(qq_cell, player.qq, player.status)
        row_base += band_height


def _render_conflict_sheet(ws: Worksheet, rows: list[dict[str, str]]) -> None:
    """渲染 QQ 冲突辅助工作表。"""
    headers = ["对局ID", "玩家昵称", "候选QQ", "采用QQ", "冲突来源条数", "最新提交时间"]
    ws.sheet_view.showGridLines = False
    for col, header in enumerate(headers, start=1):
        _style_cell(
            ws.cell(1, col),
            value=header,
            font=WHITE_FONT,
            fill=HEADER_FILL,
            border=ALL_BORDER,
            alignment=CENTER,
        )
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            _style_cell(
                ws.cell(row_idx, col_idx),
                value=row_data.get(header, ""),
                font=DATA_FONT,
                border=ALL_BORDER,
                alignment=CENTER,
            )
    widths = {1: 38.0, 2: 20.0, 3: 30.0, 4: 16.0, 5: 12.0, 6: 22.0}
    _set_column_widths(ws, widths)


def sanitize_filename_part(value: str | None) -> str:
    """清洗文件名片段，移除 Windows 非法字符。"""
    if not value:
        return ""
    cleaned = re.sub(r'[<>:"/\\\\|?*]+', "_", str(value).strip())
    return cleaned.strip(" ._")


def build_participant_list_filename(
    *,
    match_id: str,
    event_name: str | None = None,
    round_name: str | None = None,
) -> str:
    """按当前规则生成单局参赛名单文件名。"""
    event_part = sanitize_filename_part(event_name)
    round_part = sanitize_filename_part(round_name)
    match_part = sanitize_filename_part(match_id) or "match"
    if event_part and round_part:
        return f"{event_part}_{round_part}_参赛者名单.xlsx"
    if event_part:
        return f"{event_part}_{match_part}_参赛者名单.xlsx"
    if round_part:
        return f"{round_part}_{match_part}_参赛者名单.xlsx"
    return f"{match_part}_参赛者名单.xlsx"


def build_batch_participant_zip_filename(event_name: str | None = None, now: datetime | None = None) -> str:
    """按当前规则生成批量 ZIP 文件名。"""
    event_part = sanitize_filename_part(event_name)
    if event_part:
        return f"{event_part}_参赛者名单.zip"
    current = now or datetime.now()
    return f"参赛者名单_{current.strftime('%Y%m%d_%H%M%S')}.zip"


def build_batch_participant_zip(
    *,
    client: PubgAPIClient,
    match_ids: list[str],
    signup_excel_bytes: bytes | None,
    signup_sheet_schema: SignupSheetSchema | None,
    signup_mode_name: str | None,
    event_name: str | None,
    round_name_map: dict[str, str],
    current_overview: MatchOverview | None = None,
    current_teams: list[TeamSummary] | None = None,
    progress_callback: Callable[[int, int, str, str], None] | None = None,
) -> BatchParticipantListResult:
    """批量拉取多局数据并生成参赛名单 ZIP。"""
    if not match_ids:
        raise ValueError("至少选择一局才能批量生成。")

    item_filenames: list[str] = []
    failed_matches: list[dict[str, str]] = []
    total_players = 0
    total_conflicts = 0
    total_missing_contacts = 0
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, mode="w", compression=ZIP_DEFLATED) as zf:
        total = len(match_ids)
        for index, match_id in enumerate(match_ids, start=1):
            if progress_callback is not None:
                progress_callback(index, total, match_id, "start")
            try:
                if current_overview and current_overview.match_id == match_id and current_teams is not None:
                    overview = current_overview
                    teams = current_teams
                else:
                    payload = client.get_match(match_id)
                    overview = build_match_overview(match_id, payload)
                    teams = extract_team_summaries(match_id, payload)

                workbook_result = build_participant_list_workbook(
                    overview=overview,
                    teams=teams,
                    signup_excel_bytes=signup_excel_bytes,
                    signup_sheet_schema=signup_sheet_schema,
                    signup_mode_name=signup_mode_name,
                    event_name=event_name,
                    round_name=round_name_map.get(match_id),
                )
                file_name = build_participant_list_filename(
                    match_id=match_id,
                    event_name=event_name,
                    round_name=round_name_map.get(match_id),
                )
                zf.writestr(file_name, workbook_result.workbook_bytes)
                item_filenames.append(file_name)
                total_players += workbook_result.total_players
                total_conflicts += workbook_result.conflict_count
                total_missing_contacts += workbook_result.missing_contact_count
            except Exception as exc:
                failed_matches.append(
                    {
                        "match_id": match_id,
                        "round_name": round_name_map.get(match_id, ""),
                        "error": str(exc),
                    }
                )
                if progress_callback is not None:
                    progress_callback(index, total, match_id, "error")
                continue

            if progress_callback is not None:
                progress_callback(index, total, match_id, "success")

    if not item_filenames:
        raise ValueError("所选对局全部生成失败，请检查 match_id、API Key 或网络请求状态。")

    return BatchParticipantListResult(
        zip_bytes=zip_buffer.getvalue(),
        zip_filename=build_batch_participant_zip_filename(event_name),
        requested_match_count=len(match_ids),
        generated_match_count=len(item_filenames),
        total_players=total_players,
        total_conflicts=total_conflicts,
        total_missing_contacts=total_missing_contacts,
        item_filenames=item_filenames,
        failed_matches=failed_matches,
    )
